from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
import time
import pandas as pd
import openpyxl  # 設定對齊時需要
from openpyxl.styles import Alignment  # 用於靠左對齊



# 設定 Chrome WebDriver
chrome_options = Options()
chrome_options.add_argument("--disable-blink-features=AutomationControlled")  # 隱藏 Selenium
chrome_options.add_argument("--start-maximized")  # 最大化視窗
chrome_options.add_argument("--disable-infobars")  # 移除資訊欄
chrome_options.add_argument("--enable-gpu-rasterization")  # 啟用 GPU 光柵化
chrome_options.add_argument("--use-gl=desktop")  # 強制使用桌面 OpenGL
chrome_options.add_argument("--enable-webgl")  # 確保 WebGL 可用
#chrome_options.add_argument("--headless=new")  # 無頭模式 (新版)
chrome_options.add_argument(f"user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")

service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=service, options=chrome_options)

# 設定等待時間
wait = WebDriverWait(driver, 10)

try:
    #  開啟 HamaStar 官網
    driver.get("https://www.hamastar.com.tw/")
    print("成功開啟網站")

    #  確保標題正確
    expected_title = "哈瑪星科技全球資訊網"
    assert expected_title in driver.title, "網站標題不符合預期！"
    print(f"標題檢查成功: {driver.title}")

    #  測試導覽列點擊「關於我們」
    about_us = wait.until(EC.element_to_be_clickable((By.LINK_TEXT, "關於我們")))
    about_us.click()
    print("成功進入『關於我們』頁面")
    time.sleep(3)

    #  測試回到首頁
    driver.back()
    print("返回首頁")
    time.sleep(2)

    #  測試「客戶實績」頁面
    contact_us = wait.until(EC.element_to_be_clickable((By.LINK_TEXT, "客戶實績")))
    contact_us.click()
    print("成功進入『客戶實績』頁面")
    time.sleep(3)

    #  測試「我們的客戶」頁面
    contact_us = wait.until(EC.element_to_be_clickable((By.LINK_TEXT, "我們的客戶")))
    contact_us.click()
    print("成功進入『我們的客戶』頁面")
    time.sleep(3)

    #  等待 figcaption 類別的元素出現
    wait.until(EC.presence_of_element_located((By.CLASS_NAME, "figcaption")))

    #  抓取所有 figcaption 類別的內容
    elements = driver.find_elements(By.CLASS_NAME, "figcaption")


 #  儲存結果
    client_data = []
    if elements:
        print(" 成功找到 figcaption 類別的元素，內容如下：")
        for index, element in enumerate(elements, start=1):
            text = element.text.strip()
            print(f"{index}. {text}")
            client_data.append([index, text])  # 儲存到 list

    else:
        print(" 未找到任何 figcaption 類別的元素！")

    #  存入 Excel，並讓客戶名稱靠左對齊
    if client_data:
        df = pd.DataFrame(client_data, columns=["編號", "客戶名稱"])
        excel_filename = "客戶資料.xlsx"
        df.to_excel(excel_filename, index=False, engine="openpyxl")

        # **設定 Excel 內容靠左對齊**
        wb = openpyxl.load_workbook(excel_filename)
        ws = wb.active

        # 遍歷 "客戶名稱" 欄位，並設置靠左對齊
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
            for cell in row:
                cell.alignment = Alignment(horizontal="left")
        
         # 遍歷 "客戶名稱" 欄位，並設置靠左對齊
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
            for cell in row:
                cell.alignment = Alignment(horizontal="left")

        # **自動調整儲存格大小**
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # 獲取欄位名稱
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)  # 增加2以避免內容過於貼近儲存格邊緣
            ws.column_dimensions[column].width = adjusted_width


        # 儲存 Excel 檔案
        wb.save(excel_filename)

        print(f" 成功將資料存入 {excel_filename}（內容已靠左對齊）")


except Exception as e:
    print(f"發生錯誤：{e}")

finally:
    driver.quit()
    print("瀏覽器已關閉")
